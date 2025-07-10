import pandas as pd
import numpy as np
import csv
import os
import datetime
import calendar
from pandas.api.types import CategoricalDtype
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

pd.set_option('display.max_colwidth', None)
pd.set_option('display.max_rows', None)

cats = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
facility_capacity = {"Helen Newman FC":80, "Noyes FC":65, "Teagle Down FC":50, "Teagle Up FC":65, "Toni Morrison FC":75}

def convert_to_int(x):
    if pd.isna(x):
        return 0
    return int(x)

def convert_to_float(x):
    if pd.isna(x):
        return 0.0
    return float(x)

# FORMAT FROM CONNECT2
def initial_transform(filename):
    input = open(filename, 'r')
    output = open("mod_"+filename, 'w')
    writer = csv.writer(output)
    reader = csv.reader(input)

    count = 0
    for r in reader:
        if(count < 13):
            count+=1
            continue
        writer.writerow(r)  
    input.close()
    output.close()
    df = pd.read_csv("mod_"+filename)
    if os.path.exists("mod_"+filename):
        os.remove("mod_"+filename)
    df.drop(['Total Count', 'Facility','Status'], axis=1, inplace=True)
    df['Date']=df['Date'].apply(lambda x: datetime.datetime.strptime(x, '%m/%d/%Y %H:%M:%S %p'))
    df['Weekday']=df['Date'].apply(lambda x: str(calendar.day_name[x.weekday()]))

    cat_type = CategoricalDtype(categories=cats,ordered=True)

    df['Weekday'] = df['Weekday'].astype(cat_type)
    df['Location']=df['Location'].apply(lambda x: x.replace("Fitness Center", "FC"))

    hours = df.drop(['Date', 'Weekday', 'Location'], axis=1).columns # select hours from data given
    df[hours] = df[hours].replace('C', np.nan).astype(float)
    return df

# FOR GRAPH NUMBER / CAPACITY
def capacity_percent_transform_df(df):
    def compute_capacity(x): #df row
        location = x['Location'].unique()[0] # invariant: this is 1 value
        hours = df.drop(['Weekday', 'Location'], axis=1).columns
        max = facility_capacity.get(location, 9999)
        return x[hours].apply(lambda a: (a/max))

    df = df.drop(['Date'], axis=1)
    df = df.groupby(by=['Location','Weekday'], observed=False).apply(lambda x: compute_capacity(x)).droplevel(2)
    df = df.groupby(by=['Location', 'Weekday'], observed=False).mean()  
    df = df.map(lambda x: convert_to_float(x))
    return df
    
def generate_counts_reports(filename, output="output.xlsx"):
    # generate dataframes
    df = initial_transform(filename)

    df_capacity_percent = capacity_percent_transform_df(df)

    df = df.drop(['Date'], axis=1)
    df = df.groupby(by=['Location', 'Weekday'], observed=False).mean()
    df = df.map(lambda x: convert_to_int(x))

    # create excel sheets
    print("saving sheets...")

    writer = pd.ExcelWriter(output)
    df.to_excel(writer, sheet_name='capacity') # TODO: pull date range from filename
    df.to_excel(writer, sheet_name='capacity_with_density')
    df_capacity_percent.to_excel(writer, sheet_name='capacity_percent')
    df_capacity_percent.to_excel(writer, sheet_name='capacity_percent_with_density')

    workbook = writer.book

    # add % 
    worksheet = writer.sheets['capacity_percent']

    letter = chr(worksheet.max_column + 96)
    number = worksheet.max_row
    end_place = str(letter) + str(number)
    range ="C2:"+end_place

    number_format = '0%'
    for x in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']: # TODO: for every column, instead of hardcoded
        for cell in worksheet[x]:
            cell.number_format = number_format

    blueFill = PatternFill(start_color='666FFF', end_color = '666FFF', fill_type='solid')
    tealFill = PatternFill(start_color='00CCFF', end_color='00CCFF', fill_type='solid')
    greenFill = PatternFill(start_color='92D050', end_color='92D050',fill_type='solid')
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orangeFill = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')
    pinkFill = PatternFill(start_color='FF5050', end_color='FF5050', fill_type='solid')
    redFill = PatternFill(start_color='D20000', end_color='D20000', fill_type='solid')

    # add color conditional formatting
    worksheet = writer.sheets["capacity_with_density"]
    worksheet.conditional_formatting.add(
        range, 
        CellIsRule(operator='between', formula=['0', '10'], stopIfTrue=True, fill=blueFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['10','20'], stopIfTrue=True, fill=tealFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['20','30'], stopIfTrue=True, fill=greenFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['30','40'], stopIfTrue=True, fill=yellowFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['40','50'], stopIfTrue=True, fill=orangeFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['50','60'], stopIfTrue=True, fill=pinkFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['60','75'], stopIfTrue=True, fill=redFill) 
    )

    worksheet = writer.sheets["capacity_percent_with_density"]
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['0', '.20'], stopIfTrue=True, fill=blueFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['.20', '.40'], stopIfTrue=True, fill=tealFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['.40', '.60'], stopIfTrue=True, fill=greenFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['60', '80'], stopIfTrue=True, fill=yellowFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['.80', '.90'], stopIfTrue=True, fill=orangeFill) 
    )
    worksheet.conditional_formatting.add(
       range, 
        CellIsRule(operator='between', formula=['.90', '.100'], stopIfTrue=True, fill=pinkFill) 
    )

    # add % 
    number_format = '0%'
    for x in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
        for cell in worksheet[x]:
            cell.number_format = number_format

    ### move sheets
    """sheet_names = ["capacity", "capacity_with_density", "capacity_percent", "capacity_percent_with_density"]

    data_dict_row = {} # row number : values from all sheets to be written to that row
    row_num = 2
    for sheet in sheet_names:
        ws = writer.sheets[sheet]
        for row in ws.iter_rows(min_row=row_num, values_only=True):
            data_dict_row[row_num] = [row]
            row_num += 1
        row_num = 2

    print(data_dict_row[2])
    workbook.create_sheet(title="combined")
    ws = writer.sheets['combined']
    row = 0
    for new_sheet_row, data_row in data_dict_row.items():
        ws.append(data_row)"""
        
    writer.close()
    print("sheets saved.")
    return True

# For testing purposes:
def run_locally(filename="Export (1).csv"):
    generate_counts_reports(filename)

#run_locally()